# excel_updater.py
import json
from openpyxl import load_workbook

EXCEL_PATH = r"C:\Users\woain\Python_AI\Project_Car\drivers_2026.xlsx"
SHEET_NAME = "Sheet1"

INDEX_COL = 1          # A열: 번호
ACCIDENT_COL = 10      # J열: 사고점수
HEADER_ROW = 1         # 1행은 헤더

def write_accident_score(index_value: str, accident_score: int,
                         excel_path: str = EXCEL_PATH,
                         sheet_name: str = SHEET_NAME):
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]

    target_row = None
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=INDEX_COL).value
        if v is None:
            continue
        if str(v).strip() == str(index_value).strip():
            target_row = r
            break

    if target_row is None:
        raise ValueError(f'엑셀에서 번호(A열)="{index_value}" 행을 찾지 못했습니다.')

    ws.cell(row=target_row, column=ACCIDENT_COL).value = int(accident_score)
    wb.save(excel_path)
    return {"index": index_value, "row": target_row, "accident_score": int(accident_score)}

def update_from_severity_json(severity_json_path: str,
                              excel_path: str = EXCEL_PATH,
                              sheet_name: str = SHEET_NAME):
    with open(severity_json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    index_value = data.get("index") or data.get("job_id")
    if index_value is None:
        raise KeyError("severity.json에 index/job_id가 없습니다.")

    severity = int(data["severity"])  # 0~4
    return write_accident_score(index_value, severity, excel_path, sheet_name)
